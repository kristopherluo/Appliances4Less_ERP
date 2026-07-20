import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from sqlalchemy.orm import Session
from database import get_db
import models, schemas

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


@router.get("/", response_model=list[schemas.ItemOut])
def list_items(store_id: int | None = None, db: Session = Depends(get_db)):
    query = db.query(models.Item)
    if store_id:
        query = query.filter(models.Item.store_id == store_id)
    return query.order_by(models.Item.name).all()


@router.post("/", response_model=schemas.ItemOut, status_code=status.HTTP_201_CREATED)
def create_item(data: schemas.ItemCreate, db: Session = Depends(get_db)):
    item = models.Item(**data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/{item_id}", response_model=schemas.ItemOut)
def update_item(item_id: int, data: schemas.ItemUpdate, db: Session = Depends(get_db)):
    item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()


@router.get("/export")
def export_items_xlsx(store_id: int | None = None, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    from fastapi.responses import StreamingResponse
    import io as _io

    query = db.query(models.Item)
    if store_id:
        query = query.filter(models.Item.store_id == store_id)
    items = query.order_by(models.Item.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["Load Number", "Models", "Description", "Serials", "MSRP", "Store Price", "Details"]
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=item.load_number or "")
        ws.cell(row=row_idx, column=2, value=item.model_number or "")
        ws.cell(row=row_idx, column=3, value=item.appliance_type or "")
        ws.cell(row=row_idx, column=4, value=item.serial_number or "")
        ws.cell(row=row_idx, column=5, value=item.sale_price)
        ws.cell(row=row_idx, column=6, value=item.cost_price)
        ws.cell(row=row_idx, column=7, value=item.name or "")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory.xlsx"},
    )


@router.post("/import", status_code=status.HTTP_201_CREATED)
async def import_items_xlsx(
    store_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Detect header row — find row with "MODELS" or "Model" in it
    header_row_idx = None
    headers = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_lower = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(k in row_lower for k in ("models", "model", "serials", "serial")):
            header_row_idx = i
            for col_idx, cell_val in enumerate(row):
                if cell_val:
                    headers[str(cell_val).strip().lower()] = col_idx
            break

    if header_row_idx is None:
        raise HTTPException(status_code=400, detail="Could not find header row with MODELS/SERIALS columns")

    def col(names: list[str]):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    idx_load   = col(["load number", "load#", "load_number"])
    idx_model  = col(["models", "model", "model number", "model#"])
    idx_desc   = col(["description", "desc", "type"])
    idx_serial = col(["serials", "serial", "serial number", "serial#"])
    idx_msrp   = col(["msrp", "retail", "retail price"])
    idx_store  = col(["store price", "store_price", "cost", "price"])
    idx_detail = col(["details", "detail", "notes"])

    if idx_model is None or idx_serial is None:
        raise HTTPException(status_code=400, detail="Required columns MODELS and SERIALS not found")

    store = db.query(models.Store).filter(models.Store.id == store_id).first()
    if not store:
        raise HTTPException(status_code=400, detail="Store not found")

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        model_val = str(row[idx_model]).strip() if idx_model is not None and row[idx_model] else None
        serial_val = str(row[idx_serial]).strip() if idx_serial is not None and row[idx_serial] else None

        if not model_val or model_val in ("None", ""):
            skipped += 1
            continue

        desc_val   = str(row[idx_desc]).strip() if idx_desc is not None and row[idx_desc] else None
        detail_val = str(row[idx_detail]).strip() if idx_detail is not None and row[idx_detail] else None
        load_val   = str(row[idx_load]).strip() if idx_load is not None and row[idx_load] else None

        def to_float(idx):
            if idx is None or row[idx] is None:
                return 0.0
            try:
                return float(row[idx])
            except (ValueError, TypeError):
                return 0.0

        msrp_val  = to_float(idx_msrp)
        store_val = to_float(idx_store)

        name = (detail_val or desc_val or model_val)[:255]

        item = models.Item(
            store_id=store_id,
            name=name,
            appliance_type=desc_val,
            model_number=model_val,
            serial_number=serial_val,
            load_number=load_val,
            sale_price=msrp_val,
            cost_price=store_val,
            location=store.name,
            is_in_stock=True,
        )
        db.add(item)
        imported += 1

    db.commit()
    return {"imported": imported, "skipped": skipped}
